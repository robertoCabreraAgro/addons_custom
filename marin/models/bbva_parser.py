import base64
import logging
import re
from datetime import datetime
from io import BytesIO

import chardet
from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

BBVA_TXT_HEADER = ["Día", "Concepto / Referencia", "cargo", "Abono", "Saldo"]

BBVA_XLSX_HEADER = [
    "FECHA",
    "DESCRIPCIÓN",
    "ABONO",
    "CARGO",
    "SALDO",
]


class BBVAParser(models.AbstractModel):
    _name = "bbva.parser"
    _description = "BBVA files parser"

    def _detect_encoding(self, data_file):
        """
        Detects the most suitable encoding for the given file by trying multiple options.

        :param data_file: File content in bytes
        :return: Valid encoding string if found, otherwise None
        """
        encodings = [
            chardet.detect(data_file)["encoding"],
            "utf-8",
            "iso-8859-1",
            "cp1252",
            "latin1",
        ]

        # Decode content and split into lines
        for encoding in encodings:
            try:
                lines = data_file.decode(encoding, errors="replace").splitlines()

                if not lines:
                    continue

                header = lines[0].strip().split("\t")

                if header == BBVA_TXT_HEADER:
                    return encoding

            except UnicodeDecodeError:
                continue

        return None

    def _validate_header(self, data_file, file_type):
        """Validates if the file header matches the expected format based on file type.

        :param data_file: Data file in bytes
        :param file_type: Type of file ('txt' or 'xlsx')
        :return: True if the header is valid, False otherwise
        """
        try:
            if file_type == "txt":
                return bool(self._detect_encoding(data_file))
            elif file_type == "xlsx":
                header_row = 6
                workbook = load_workbook(
                    BytesIO(data_file), read_only=True, data_only=True
                )
                sheet = workbook.active

                if sheet.max_row <= header_row:
                    return False

                header = [
                    cell.value.strip() if cell.value else ""
                    for cell in sheet[header_row]
                ]

                return header == BBVA_XLSX_HEADER

            return False

        except Exception as e:
            _logger.error("Error validating %s header: %s", file_type, str(e))
            return False

    def _generate_unique_import_id(
        self, journal_code, transaction_date, sequence, padding=4
    ):
        """
        Generates a universal unique import ID with standardized format.

        Format: {JOURNAL_CODE}-{YYMM}-{SEQ}

        Args:
            journal_code (str): Journal's short code (e.g., 'BNK-BBVA')
            transaction_date (date): Date object or date string (YYYY-MM-DD)
            sequence (int): Transaction sequence number
            padding (int): Zero-padding length for sequence (default: 4)

        Returns:
            str: Formatted unique import ID (e.g., 'BNK-BBVA-2403-0042')
        """
        try:
            # Ensure journal code is clean (no special chars)
            clean_journal_code = re.sub(r"[^A-Z0-9-]", "", journal_code.upper().strip())

            # Parse date (handles both string and date object)
            if isinstance(transaction_date, str):
                dt = fields.Date.from_string(transaction_date)
            else:
                dt = transaction_date

            # Format components
            yymm = dt.strftime("%y%m")  # Last 2 digits of year + 2-digit month
            seq_str = str(sequence).zfill(padding)  # Zero-padded sequence

            return f"{clean_journal_code}-{yymm}-{seq_str}"

        except Exception as e:
            raise ValueError(
                _("Failed to generate transaction ID. Please check input parameters.")
            )

    def _parse_date(self, date_str, date_format):
        """Parse date from string using specified format.

        Args:
            date_str (str): Date string
            date_format (str): Format string for datetime.strptime

        Returns:
            str: Date string in Odoo format

        Raises:
            UserError: If date format is invalid
        """
        try:
            date_obj = datetime.strptime(date_str, date_format)
            return fields.Date.to_string(date_obj)
        except ValueError:
            raise UserError(
                _("Invalid date format: %s. Expected format: %s")
                % (date_str, date_format)
            )

    def _find_partner_bank(self, reference):
        """Try to find partner based on reference (description).

        Args:
            reference (str): Transaction reference/description

        Returns:
            res.partner.bank: Partner bank record if found, empty recordset otherwise
        """
        bank_acc_obj = self.env["res.partner.bank"]

        # Buscar números de cuenta en la referencia
        account_numbers = re.findall(r"\b\d{10,20}\b", reference)
        for acc_num in account_numbers:
            bank_account = bank_acc_obj.search(
                [("acc_number", "ilike", acc_num)], limit=1
            )
            if bank_account:
                return bank_account

        return bank_acc_obj

    def _validate_date_range(self, date_start, date_end):
        """Validate that dates are from the same month and year.

        Args:
            date_start (str): Start date in Odoo format
            date_end (str): End date in Odoo format

        Raises:
            UserError: If dates are from different months or years
        """

        if not date_start or not date_end:
            raise UserError(_("No valid transactions found in the file"))

        start_date = fields.Date.to_date(date_start)
        end_date = fields.Date.to_date(date_end)

        if start_date.year != end_date.year or start_date.month != end_date.month:
            raise UserError(
                _(
                    "All transactions must be from the same month and year. "
                    "Start date: %(start)s, End date: %(end)s"
                )
                % {"start": start_date, "end": end_date}
            )

    def parse_bbva_file(self, data_file, journal, file_type):
        """Parse BBVA file (TXT or XLSX format).

        Args:
            data_file (bytes): The file content to parse
            journal (account.journal): Journal record
            file_type (str): File type ('txt' or 'xlsx')

        Returns:
            dict: Contains statement name, date, and transactions

        Raises:
            UserError: If file format is invalid or contains inconsistent data
        """
        if not self._validate_header(data_file, file_type):
            raise UserError(
                _("Invalid file format: Header does not match expected BBVA format")
            )

        try:
            if file_type == "txt":
                return self._parse_txt_file(data_file, journal)
            elif file_type == "xlsx":
                return self._parse_xlsx_file(data_file, journal)
            else:
                raise UserError(
                    _("Unsupported file type. Only TXT and XLSX formats are supported.")
                )

        except UnicodeDecodeError:
            raise UserError(_("Encoding error: Could not decode the file content"))
        except Exception as e:
            raise UserError(_("File processing error:\n\t- %s") % str(e))

    def _parse_txt_file(self, data_file, journal):
        """Parse BBVA TXT file format.

        Args:
            data_file (bytes): The file content to parse
            journal (account.journal): Journal record

        Returns:
            dict: Contains statement name, date, and transactions
        """
        # Decode file and convert it into lines
        encoding_detected = self._detect_encoding(data_file)
        lines = data_file.decode(encoding_detected, errors="replace").splitlines()

        if not lines:
            raise UserError(_("Empty file: The uploaded file contains no data"))

        transactions = []
        date_start = None
        date_end = None
        sequence = 1
        lines_number = len(lines)
        for line in reversed(lines[1:]):
            # Skip empty lines
            if not line.strip():
                continue

            # Validate and process each line
            values = line.strip().split("\t")
            if len(values) != 5:
                raise UserError(
                    _("Line %d: Invalid format, expected 5 columns")
                    % (lines_number - sequence)
                )

            # Get cell values
            try:
                day_str = values[0].strip()
                date = self._parse_date(day_str, "%d-%m-%Y")
                reference = values[1].strip()
                debit = float(values[2].replace(",", "").strip() or "0")
                credit = float(values[3].replace(",", "").strip() or "0")
            except (ValueError, IndexError) as e:
                raise UserError(
                    _("Line %d: Invalid data format - %s")
                    % ((lines_number - sequence), str(e))
                )

            # Validate required values
            if not (debit or credit):
                raise UserError(
                    _("Line %d: Transaction must have either debit or credit amount")
                    % (lines_number - sequence)
                )

            # Set start/end dates
            if date_end is None:
                date_end = date
            date_start = date  # Always updates with last valid date

            # Create transaction
            partner_bank = self._find_partner_bank(reference)
            transactions.append(
                {
                    "date": date,
                    "payment_ref": reference,
                    "amount": credit - debit,
                    "account_number": partner_bank.acc_number if partner_bank else None,
                    "partner_id": partner_bank.partner_id.id
                    if partner_bank.partner_id
                    else None,
                    "sequence": 1,  # always 1 to be consistent with running balance computing
                    "unique_import_id": self._generate_unique_import_id(
                        journal.code, date, sequence
                    ),
                }
            )

            # Update sequence
            sequence += 1

        # Final validations
        self._validate_date_range(date_start, date_end)

        # Create statement data
        statement_name = fields.Date.from_string(date_start).strftime("%y-%m")
        return {
            "name": statement_name,  # YY-MM format from start date
            "date": date_end,  # Using last transaction date as statement date
            "transactions": transactions,
        }

    def _parse_xlsx_file(self, data_file, journal):
        """Parse BBVA XLSX file format.

        Args:
            data_file (bytes): The Excel file content to parse
            journal (account.journal): Journal record

        Returns:
            dict: Contains statement name, date, and transactions
        """
        try:
            workbook = load_workbook(BytesIO(data_file), read_only=True, data_only=True)
            sheet = workbook.active
            header_row = 6

            # Check if exists at least one row after header
            if sheet.max_row <= header_row + 1:
                raise UserError(
                    _("Empty file: The Excel file contains no transaction data")
                )

            transactions = []
            date_start = None
            date_end = None
            sequence = 1
            footer_row = sheet.max_row
            # Iterate from the last row upwards
            for row_number in range(footer_row - 1, header_row, -1):
                # Get row as a list
                row = [cell.value for cell in list(sheet.rows)[row_number - 1]]

                # Skip empty rows
                if not any(row):
                    continue

                # Get cell values
                try:
                    date_value = row[0]
                    date = (
                        fields.Date.to_string(date_value)
                        if isinstance(date_value, datetime)
                        else self._parse_date(date_value, "%Y-%m-%d")
                    )
                    reference = str(row[1] or "").strip()
                    credit = float(row[2] or 0)
                    debit = float(row[3] or 0)
                except (ValueError, IndexError) as e:
                    raise UserError(
                        _("Row %d: Invalid data format - %s") % (row_number, str(e))
                    )

                # Validate required values
                if not (debit or credit):
                    raise UserError(
                        _("Row %d: Transaction must have either debit or credit amount")
                        % (row_number)
                    )

                # Set start/end dates
                if date_end is None:
                    date_end = date
                date_start = date

                # Create transaction
                partner_bank = self._find_partner_bank(reference)
                transactions.append(
                    {
                        "date": date,
                        "payment_ref": reference,
                        "amount": credit - debit,
                        "account_number": partner_bank.acc_number
                        if partner_bank
                        else None,
                        "partner_id": partner_bank.partner_id.id
                        if partner_bank.partner_id
                        else None,
                        "sequence": 1,  # always 1 to be consistent with running balance computing
                        "unique_import_id": self._generate_unique_import_id(
                            journal.code, date, sequence
                        ),
                    }
                )

                # Update sequence
                sequence += 1

            # Final validations
            self._validate_date_range(date_start, date_end)

            # Create statement data
            statement_name = fields.Date.from_string(date_start).strftime("%y-%m")
            return {
                "name": statement_name,  # YY-MM format from start date
                "date": date_end,  # Using last transaction date as statement date
                "transactions": transactions,
            }

        except Exception as e:
            raise UserError(_("Excel file error: %s") % str(e))
